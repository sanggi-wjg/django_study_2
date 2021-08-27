import os
from typing import List

import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from django.views.generic.base import View
from openpyxl import load_workbook

from amk_demo.settings.base import MEDIA_ROOT
from apps.modules.helpers.upload_file_helper import upload_file_to_server
from apps.modules.mixins.auth_mixins import LoginRequired
from apps.sample.forms import UploadFileForm


class SampleController(LoginRequired, View):
    view_title = '샘플'
    template_name = 'sample/sample_content.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, context = {
            'view_title' : self.view_title,
            'upload_form': UploadFileForm()
        })

    def post(self, request, *args, **kwargs):
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            upload_file_to_server(request.FILES.get('uploadFile'))


class SampleMultiController(LoginRequired, View):
    view_title = '샘플'
    template_name = 'sample/sample_multi_content.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, context = {
            'view_title' : self.view_title,
            'upload_form': UploadFileForm()
        })


class SampleInsertExcelController(LoginRequired, View):

    def get(self, request, *args, **kwargs):
        filepath = os.path.join(MEDIA_ROOT, 'sample_excel_2.xlsx')
        service = ReadExcelSample(filepath)
        service.operate()

        return HttpResponse('Success')


########################################################################################################################

class DataSampleVO(object):

    def __init__(self, index: str, rack_code: str, product_code: str, register_date: str):
        self._index = index
        self._rack_code = rack_code
        self._product_code = product_code
        self._register_date = register_date

    def validate(self):
        pass

    @property
    def index(self):
        return self._index

    @property
    def rack_code(self):
        return self._rack_code

    @property
    def product_code(self):
        return self._product_code

    @property
    def register_date(self):
        return self._register_date


class ReadExcelSampleAbs(object):
    _workbook: openpyxl.workbook.Workbook

    def __init__(self, filepath: str):
        self._filepath: str = filepath

    def _read_sheet_contents(self):
        raise NotImplementedError('_read_sheet_contents is not implemented')

    def _get_sheet_by_name(self, sheet_name: str):
        return self._workbook[sheet_name]

    def _get_sheet_by_index(self, index: int):
        return self._workbook.worksheets[index]

    def _get_sheet_names(self):
        return self._workbook.sheetnames

    def operate(self):
        self._workbook = load_workbook(filename = self._filepath)
        try:
            datasets = self._read_sheet_contents()
        finally:
            self._workbook.close()

        for d in datasets:
            print(d.index, d.rack_code, d.product_code, d.register_date)


class ReadExcelSample(ReadExcelSampleAbs):
    _skip_sheet_index = [1]
    _min_row = 2

    def _read_sheet_contents(self) -> List[DataSampleVO]:
        result = []
        for sheet_no, sheet_name in enumerate(self._get_sheet_names()):
            rows = self._workbook[sheet_name].iter_rows(min_row = self._min_row)

            for row in rows:
                result.append(self._read_rows(row))

        return result

    def _read_rows(self, row) -> DataSampleVO:
        index, rack_code, product_code, register_date = '', '', '', ''

        for cell in row:
            # print(cell.column_letter, cell.col_idx, cell.value)
            if cell.column_letter == 'A':
                index = cell.value
            if cell.column_letter == 'B':
                rack_code = cell.value
            if cell.column_letter == 'F':
                register_date = cell.value
            if cell.column_letter == 'I':
                product_code = cell.value

        vo = DataSampleVO(index, rack_code, product_code, register_date)
        vo.validate()
        return vo
