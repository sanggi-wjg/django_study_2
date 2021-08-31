import openpyxl
from openpyxl import load_workbook


class ReadExcelSampleAbs(object):
    _workbook: openpyxl.workbook.Workbook

    def __init__(self, filepath: str):
        self._filepath: str = filepath

    def _read_sheet_contents(self):
        raise NotImplementedError('_read_sheet_contents is not implemented')

    def _get_sheet_by_name(self, sheet_name: str):
        return self._workbook[sheet_name]

    def _get_sheet_by_index(self, index: int):
        return self._workbook.worksheets[index]

    def _get_sheet_names(self):
        return self._workbook.sheetnames

    def operate(self):
        self._workbook = load_workbook(filename = self._filepath)
        try:
            datasets = self._read_sheet_contents()
        finally:
            self._workbook.close()

        for d in datasets:
            print(d.index, d.rack_code, d.product_code, d.register_date)
